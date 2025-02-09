# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

sheet_name = 'Dados'  # Nome para a planilha

workbook = load_workbook(WORKBOOK_PATH)  # Carregando um arquivo do excel

# worksheet = workbook.active
worksheet = workbook[sheet_name]  # Seleciona a planilha
assert isinstance(worksheet, Worksheet)  # Garante que seja do tipo Worksheet

for row in worksheet.iter_rows(min_row=2):
    for col in row:
        print(col.value, end='\t')

        if col.value == 'Maria':
            worksheet.cell(col.row, 2, 23)  # type: ignore

    print()

# worksheet['B3'].value = '14'

workbook.save(WORKBOOK_PATH)
